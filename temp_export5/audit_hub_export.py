from __future__ import annotations

from collections import Counter
from pathlib import Path
import json

from openpyxl import load_workbook

WORKBOOK = Path(r"c:\Users\bkaufman\Adops Intelligence Hub\Adops Intelligence Hub (10).xlsx")

EXPECTED_HEADERS = {
    "Raw_Imported_Events": [
        "Event Date",
        "Source Project",
        "Network Name",
        "Advertiser",
        "Placement ID",
        "Issue Type",
        "Issue Flags",
        "Account REP OPS",
    ],
    "Normalized_Event_Ledger": [
        "Event Date",
        "Source Project",
        "Network Name",
        "Advertiser",
        "Placement ID",
        "Issue Type",
        "Issue Flags",
        "Account REP OPS",
    ],
    "Summary_By_System": ["Source Project", "Issue Count"],
    "Summary_By_Network": ["Network Name", "Issue Count"],
    "Summary_By_Issue_Type": ["Issue Flags", "Issue Count"],
    "Trend_Weekly": ["Event Week", "Source Project", "Issue Count"],
    "Trend_Monthly": ["Event Month", "Source Project", "Issue Count"],
    "Network_Grading": ["Network Name"],
    "Rep_Grading": ["AdOps Rep Performance Grading"],
    "Executive_Snapshot": ["Section", "Metric", "Value", "Status"],
    "Presentation_View": ["Leadership Snapshot"],
    "Run_Log": ["Timestamp", "Action", "Status", "Message", "Context"],
    "CVI_Daily_Baseline": ["Snapshot Date", "Placement ID", "Advertiser"],
}

CRITICAL_BLANK_CHECKS = {
    "Raw_Imported_Events": [
        "Event Date",
        "Source Project",
        "Placement ID",
        "Issue Type",
    ],
    "Normalized_Event_Ledger": [
        "Event Date",
        "Source Project",
        "Placement ID",
        "Issue Type",
        "Account REP OPS",
    ],
    "CVI_Daily_Baseline": [
        "Snapshot Date",
        "Placement ID",
        "Advertiser",
    ],
}


def is_non_empty(v):
    return not (v is None or (isinstance(v, str) and v.strip() == ""))


def row_has_data(row):
    return any(is_non_empty(v) for v in row)


def normalize_header(v):
    if v is None:
        return ""
    return str(v).strip()


def main():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)

    report = {
        "workbook": str(WORKBOOK.name),
        "sheets": {},
        "findings": [],
        "cross_checks": {},
    }

    sheet_tables = {}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            headers = []
            data_rows = []
        else:
            headers = [normalize_header(v) for v in rows[0]]
            data_rows = rows[1:]

        non_empty_rows = [r for r in data_rows if row_has_data(r)]
        non_empty_count = len(non_empty_rows)

        width = 0
        for row in rows:
            for idx, v in enumerate(row, start=1):
                if is_non_empty(v) and idx > width:
                    width = idx

        sheet_info = {
            "headers": headers,
            "header_count": len([h for h in headers if h]),
            "non_empty_rows": non_empty_count,
            "used_columns": width,
        }

        report["sheets"][ws.title] = sheet_info

        if ws.title in EXPECTED_HEADERS:
            missing_headers = [h for h in EXPECTED_HEADERS[ws.title] if h not in headers]
            if missing_headers:
                report["findings"].append(
                    {
                        "severity": "high",
                        "sheet": ws.title,
                        "type": "missing_headers",
                        "details": missing_headers,
                    }
                )

        if ws.title in CRITICAL_BLANK_CHECKS and headers:
            h_index = {h: i for i, h in enumerate(headers)}
            for field in CRITICAL_BLANK_CHECKS[ws.title]:
                if field not in h_index:
                    continue
                idx = h_index[field]
                blank = 0
                total = 0
                for row in non_empty_rows:
                    total += 1
                    val = row[idx] if idx < len(row) else None
                    if not is_non_empty(val):
                        blank += 1
                if total > 0 and blank > 0:
                    report["findings"].append(
                        {
                            "severity": "medium" if blank / total < 0.05 else "high",
                            "sheet": ws.title,
                            "type": "blank_critical_field",
                            "field": field,
                            "blank": blank,
                            "total": total,
                            "pct": round((blank / total) * 100, 2),
                        }
                    )

        if headers:
            table = []
            for row in non_empty_rows:
                obj = {}
                for i, h in enumerate(headers):
                    if not h:
                        continue
                    obj[h] = row[i] if i < len(row) else None
                table.append(obj)
            sheet_tables[ws.title] = table

    # Cross-checks
    norm = sheet_tables.get("Normalized_Event_Ledger", [])
    report["cross_checks"]["normalized_rows"] = len(norm)

    for s_name, key_field in [
        ("Summary_By_System", "Source Project"),
        ("Summary_By_Network", "Network Name"),
        ("Summary_By_Issue_Type", "Issue Flags"),
    ]:
        tab = sheet_tables.get(s_name, [])
        total = 0
        for r in tab:
            try:
                total += int(r.get("Issue Count") or 0)
            except Exception:
                pass
        report["cross_checks"][f"{s_name}_sum_issue_count"] = total
        if norm and total != len(norm):
            report["findings"].append(
                {
                    "severity": "high",
                    "sheet": s_name,
                    "type": "count_mismatch_vs_normalized",
                    "summary_total": total,
                    "normalized_rows": len(norm),
                }
            )

    # Trend table emptiness check
    for s_name in ["Trend_Weekly", "Trend_Monthly"]:
        tab = sheet_tables.get(s_name, [])
        if norm and len(tab) == 0:
            report["findings"].append(
                {
                    "severity": "medium",
                    "sheet": s_name,
                    "type": "empty_table",
                    "details": "Trend tab is empty while normalized data exists",
                }
            )

    # Run log error scan
    run_log = sheet_tables.get("Run_Log", [])
    if run_log:
        err_count = 0
        latest_errors = []
        for r in run_log:
            status = str(r.get("Status") or "").upper()
            if status == "ERROR":
                err_count += 1
                if len(latest_errors) < 5:
                    latest_errors.append(
                        {
                            "timestamp": r.get("Timestamp"),
                            "action": r.get("Action"),
                            "message": r.get("Message"),
                        }
                    )
        report["cross_checks"]["run_log_error_rows"] = err_count
        if err_count > 0:
            report["findings"].append(
                {
                    "severity": "medium",
                    "sheet": "Run_Log",
                    "type": "error_rows_present",
                    "count": err_count,
                    "sample": latest_errors,
                }
            )

    # Sort findings by severity
    rank = {"high": 0, "medium": 1, "low": 2}
    report["findings"].sort(key=lambda x: rank.get(x.get("severity", "low"), 9))

    print(json.dumps(report, default=str, indent=2))


if __name__ == "__main__":
    main()
