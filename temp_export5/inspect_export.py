from pathlib import Path
from openpyxl import load_workbook

path = max(Path('.').glob('Adops Intelligence Hub (*.xlsx'), key=lambda p: p.stat().st_mtime)
wb = load_workbook(path, data_only=True, read_only=True)

for sheet_name in ['Run_Log', 'Summary_By_System', 'Executive_Snapshot', 'Presentation_View']:
    print('\n' + '=' * 40)
    print(sheet_name)
    print('=' * 40)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[-20:]:
        print(row)
