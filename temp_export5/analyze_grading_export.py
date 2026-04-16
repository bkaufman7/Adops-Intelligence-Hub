from pathlib import Path
from statistics import median
from openpyxl import load_workbook


def latest_workbook():
    files = list(Path('.').glob('Adops Intelligence Hub (*.xlsx'))
    if not files:
        raise FileNotFoundError('No workbook export found')
    return max(files, key=lambda p: p.stat().st_mtime)


def normalize(v):
    if v is None:
        return ''
    return str(v).strip()


def to_percent(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        # Some exports store as 0.047 while others store as 4.7
        return float(v) * 100 if float(v) <= 1 else float(v)
    s = str(v).replace('%', '').strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def read_table(ws):
    rows = [r for r in ws.iter_rows(values_only=True)]
    if not rows:
        return [], []
    headers = [normalize(h) for h in rows[0]]
    data = []
    for row in rows[1:]:
        if not any(c not in (None, '') for c in row):
            continue
        data.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers)) if headers[i]})
    return headers, data


def grade_summary(data, name_field):
    grade_counts = {}
    pct_values = []
    low_denom_high_pct = []
    top = []

    for row in data:
        grade = normalize(row.get('Grade') or row.get('Issue Density Grade (Diagnostic)') or 'N/A')
        grade_counts[grade] = grade_counts.get(grade, 0) + 1

        pct = to_percent(row.get('Flagged %'))
        if pct is not None:
            pct_values.append(pct)
            top.append((pct, normalize(row.get(name_field)), int(row.get('Total Live Placements') or 0), int(row.get('Flagged Placements') or 0), int(row.get('Flagged/Live Ratio') or 0) if isinstance(row.get('Flagged/Live Ratio'), (int, float)) else None))

            total_live = int(row.get('Total Live Placements') or 0)
            flagged_live = int(row.get('Flagged Placements') or 0)
            if total_live > 0 and total_live < 25 and pct >= 20:
                low_denom_high_pct.append((normalize(row.get(name_field)), pct, flagged_live, total_live))

    top.sort(reverse=True, key=lambda x: x[0])

    f_low_denom = {'<=5': 0, '<=10': 0, '<=25': 0, 'all_f': 0}
    for row in data:
        grade = normalize(row.get('Grade') or row.get('Issue Density Grade (Diagnostic)') or 'N/A')
        if grade != 'F':
            continue
        total_live = int(row.get('Total Live Placements') or 0)
        f_low_denom['all_f'] += 1
        if total_live <= 5:
            f_low_denom['<=5'] += 1
        if total_live <= 10:
            f_low_denom['<=10'] += 1
        if total_live <= 25:
            f_low_denom['<=25'] += 1

    return {
        'grade_counts': grade_counts,
        'pct_median': median(pct_values) if pct_values else None,
        'pct_max': max(pct_values) if pct_values else None,
        'pct_min': min(pct_values) if pct_values else None,
        'top5': top[:5],
        'low_denom_high_pct': low_denom_high_pct,
        'f_low_denom': f_low_denom,
    }


def main():
    wb_path = latest_workbook()
    wb = load_workbook(wb_path, data_only=True, read_only=True)

    print(f'Workbook: {wb_path.name}')

    # Executive snapshot KPI
    if 'Executive_Snapshot' in wb.sheetnames:
        headers, rows = read_table(wb['Executive_Snapshot'])
        kpi = None
        for row in rows:
            if normalize(row.get('Metric')) == 'Primary Grading KPI: Flagged % of Live Placements':
                kpi = to_percent(row.get('Value'))
                break
        print(f'Executive KPI (Flagged % of Live): {kpi:.2f}%' if kpi is not None else 'Executive KPI not found')

    if 'Rep_Grading' in wb.sheetnames:
        _, rep_rows = read_table(wb['Rep_Grading'])
        rep_summary = grade_summary(rep_rows, 'Rep')
        print('\nRep_Grading Summary')
        print('Grade counts:', rep_summary['grade_counts'])
        print('Pct range:', rep_summary['pct_min'], 'to', rep_summary['pct_max'])
        print('Pct median:', rep_summary['pct_median'])
        print('Top 5 by Flagged %:')
        for item in rep_summary['top5']:
            print(' ', item)
        print('Low denominator high pct count (<25 live and >=20%):', len(rep_summary['low_denom_high_pct']))
        print('F denominator breakdown:', rep_summary['f_low_denom'])

    if 'Advertiser_Grading' in wb.sheetnames:
        _, adv_rows = read_table(wb['Advertiser_Grading'])
        adv_summary = grade_summary(adv_rows, 'Advertiser')
        print('\nAdvertiser_Grading Summary')
        print('Grade counts:', adv_summary['grade_counts'])
        print('Pct range:', adv_summary['pct_min'], 'to', adv_summary['pct_max'])
        print('Pct median:', adv_summary['pct_median'])
        print('Top 5 by Flagged %:')
        for item in adv_summary['top5']:
            print(' ', item)
        print('F denominator breakdown:', adv_summary['f_low_denom'])

        archive_rows = [r for r in adv_rows if normalize(r.get('Advertiser')).lower().startswith('xarchive_')]
        print('Archive-prefixed advertisers in table:', len(archive_rows))
        if archive_rows:
            pcts = [to_percent(r.get('Flagged %')) for r in archive_rows if to_percent(r.get('Flagged %')) is not None]
            print('Archive advertisers pct range:', (min(pcts) if pcts else None), 'to', (max(pcts) if pcts else None))


if __name__ == '__main__':
    main()
