from collections import Counter, defaultdict
from pathlib import Path
import sys

from openpyxl import load_workbook

WORKBOOK_GLOB = "Adops Intelligence Hub (*.xlsx"

try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

RAW_REQUIRED = [
    'Event Date',
    'Source Project',
    'Network Name',
    'Advertiser',
    'Placement ID',
    'Placement Name',
    'Issue Type',
    'Issue Flags',
    'Account REP OPS',
    'Source File Name',
    'Export Timestamp',
]
NORMALIZED_REQUIRED = [
    'Event Date',
    'Source Project',
    'Network Name',
    'Advertiser',
    'Placement ID',
    'Issue Type',
    'Issue Flags',
    'Account REP OPS',
]
BASELINE_REQUIRED = [
    'Snapshot Date',
    'Network ID',
    'Advertiser',
    'Placement ID',
    'Placement',
]
def pick_latest_workbook():
    files = sorted(Path('.').glob('Adops Intelligence Hub (*.xlsx'))
    if not files:
        raise FileNotFoundError('No Adops Intelligence Hub export found in workspace root.')
    return max(files, key=lambda path: path.stat().st_mtime)
def normalize_header(value):
    return str(value).strip() if value is not None else ''


def is_blank(value):
    return value is None or (isinstance(value, str) and value.strip() == '')
def is_blank_row(row):
    return not row or all(is_blank(v) for v in row)


def read_sheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [normalize_header(h) for h in rows[0]]
    return headers, rows[1:]


def summarize_sheet(ws):
    headers, rows = read_sheet_rows(ws)
    nonblank_rows = [row for row in rows if not is_blank_row(row)]
    return {
        'headers': headers,
        'row_count': len(nonblank_rows),
        'col_count': len(headers),
        'rows': nonblank_rows,
    }


def header_index(headers):
    return {header: idx for idx, header in enumerate(headers)}


def missing_header_report(headers, required):
    idx = header_index(headers)
    return [col for col in required if col not in idx]
def completeness_report(headers, rows, required, group_field=None):
    idx = header_index(headers)
    by_group = defaultdict(lambda: {'rows': 0, 'missing': Counter()})

    for row in rows:
        group_value = 'ALL'
        if group_field and group_field in idx:
            raw_group = row[idx[group_field]]
            group_value = str(raw_group).strip() if raw_group is not None else '(blank)'
            if not group_value:
                group_value = '(blank)'
        target = by_group[group_value]
        target['rows'] += 1
        for col in required:
            if col not in idx:
                continue
            if is_blank(row[idx[col]]):
                target['missing'][col] += 1

    return by_group
def unique_nonblank_count(headers, rows, field):
    idx = header_index(headers)
    if field not in idx:
        return 0
    values = set()
    for row in rows:
        value = row[idx[field]]
        if not is_blank(value):
            values.add(str(value).strip())
    return len(values)
def count_data_rows(ws):
    headers, rows = read_sheet_rows(ws)
    return headers, [row for row in rows if not is_blank_row(row)]


def print_section(title):
    print('\n' + '=' * 80)
    print(title)
    print('=' * 80)


def print_sheet_inventory(wb):
    print_section('WORKBOOK INVENTORY')
    for name in wb.sheetnames:
        ws = wb[name]
        summary = summarize_sheet(ws)
        print(f"{name}: rows={summary['row_count']}, cols={summary['col_count']}")
def audit_raw_imported_events(wb):
    name = 'Raw_Imported_Events'
    if name not in wb.sheetnames:
        print_section(name)
        print('MISSING SHEET')
        return

    ws = wb[name]
    headers, rows = count_data_rows(ws)
    idx = header_index(headers)

    print_section(name)
    print('Rows:', len(rows))
    missing_headers = missing_header_report(headers, RAW_REQUIRED)
    print('Missing headers:', missing_headers or 'None')

    grouped = completeness_report(headers, rows, RAW_REQUIRED, 'Source Project')
    for source in sorted(grouped):
        info = grouped[source]
        print(f"\n[{source}] rows={info['rows']}")
        for col, count in sorted(info['missing'].items(), key=lambda item: (-item[1], item[0])):
            pct = (count / info['rows']) * 100 if info['rows'] else 0
            print(f"  missing {col}: {count} ({pct:.1f}%)")
    if 'Placement ID' in idx:
        placement_count = unique_nonblank_count(headers, rows, 'Placement ID')
        print('\nUnique placement IDs:', placement_count)


def audit_normalized_ledger(wb):
    name = 'Normalized_Event_Ledger'
    if name not in wb.sheetnames:
        print_section(name)
        print('MISSING SHEET')
        return
    ws = wb[name]
    headers, rows = count_data_rows(ws)
    print_section(name)
    print('Rows:', len(rows))
    missing_headers = missing_header_report(headers, NORMALIZED_REQUIRED)
    print('Missing headers:', missing_headers or 'None')

    grouped = completeness_report(headers, rows, NORMALIZED_REQUIRED, 'Source Project')
    for source in sorted(grouped):
        info = grouped[source]
        print(f"\n[{source}] rows={info['rows']}")
        for col, count in sorted(info['missing'].items(), key=lambda item: (-item[1], item[0])):
            pct = (count / info['rows']) * 100 if info['rows'] else 0
            print(f"  missing {col}: {count} ({pct:.1f}%)")


def audit_baseline(wb):
    name = 'CVI_Daily_Baseline'
    if name not in wb.sheetnames:
        print_section(name)
        print('MISSING SHEET')
        return
    ws = wb[name]
    headers, rows = count_data_rows(ws)
    print_section(name)
    print('Rows:', len(rows))
    missing_headers = missing_header_report(headers, BASELINE_REQUIRED)
    print('Missing headers:', missing_headers or 'None')

    grouped = completeness_report(headers, rows, BASELINE_REQUIRED, 'Snapshot Date')
    snapshot_dates = sorted(grouped)
    print('Snapshot dates:', ', '.join(snapshot_dates[-7:]) if snapshot_dates else 'None')
    if snapshot_dates:
        latest = snapshot_dates[-1]
        info = grouped[latest]
        print(f"Latest snapshot {latest}: rows={info['rows']}")
        for col, count in sorted(info['missing'].items(), key=lambda item: (-item[1], item[0])):
            pct = (count / info['rows']) * 100 if info['rows'] else 0
            print(f"  missing {col}: {count} ({pct:.1f}%)")


def audit_simple_table(wb, sheet_name):
    print_section(sheet_name)
    if sheet_name not in wb.sheetnames:
        print('MISSING SHEET')
        return

    ws = wb[sheet_name]
    headers, rows = count_data_rows(ws)
    print('Rows:', len(rows))
    print('Headers:', headers[:12])
    if rows:
        print('First data row:', [rows[0][i] for i in range(min(len(headers), 8))])
    else:
        print('NO DATA ROWS')


def audit_rep_grading(wb):
    name = 'Rep_Grading'
    print_section(name)
    if name not in wb.sheetnames:
        print('MISSING SHEET')
        return
    ws = wb[name]
    _, rows = count_data_rows(ws)
    print('Rows:', len(rows))
    rep_lines = 0
    ratio_lines = 0
    for row in rows:
        value = '' if not row else str(row[0] or '')
        if '[Grade:' in value:
            rep_lines += 1
        if 'Live Placement Ratio' in value:
            ratio_lines += 1
    print('Rep entries:', rep_lines)
    print('Ratio lines:', ratio_lines)


def audit_network_grading(wb):
    name = 'Network_Grading'
    print_section(name)
    if name not in wb.sheetnames:
        print('MISSING SHEET')
        return
    ws = wb[name]
    _, rows = count_data_rows(ws)
    print('Rows:', len(rows))
    network_lines = 0
    for row in rows:
        value = '' if not row else str(row[0] or '')
        if '[Grade:' in value:
            network_lines += 1
    print('Network entries:', network_lines)


def audit_snapshot_tabs(wb):
    for sheet_name in ['Executive_Snapshot', 'Presentation_View', 'Summary_By_System', 'Summary_By_Network', 'Summary_By_Issue_Type', 'Trend_Weekly', 'Trend_Monthly', 'Unmapped_Networks', 'Run_Log']:
        audit_simple_table(wb, sheet_name)


def cross_check_counts(wb):
    print_section('CROSS-CHECK COUNTS')
    raw_rows = summarize_sheet(wb['Raw_Imported_Events'])['row_count'] if 'Raw_Imported_Events' in wb.sheetnames else 0
    normalized_rows = summarize_sheet(wb['Normalized_Event_Ledger'])['row_count'] if 'Normalized_Event_Ledger' in wb.sheetnames else 0
    baseline_rows = summarize_sheet(wb['CVI_Daily_Baseline'])['row_count'] if 'CVI_Daily_Baseline' in wb.sheetnames else 0
    print('Raw_Imported_Events rows:', raw_rows)
    print('Normalized_Event_Ledger rows:', normalized_rows)
    print('CVI_Daily_Baseline rows:', baseline_rows)
    print('Raw minus Normalized:', raw_rows - normalized_rows)


def main():
    workbook_path = pick_latest_workbook()
    print('Workbook:', workbook_path.name)
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    print_sheet_inventory(wb)
    audit_raw_imported_events(wb)
    audit_normalized_ledger(wb)
    audit_baseline(wb)
    audit_rep_grading(wb)
    audit_network_grading(wb)
    audit_snapshot_tabs(wb)
    cross_check_counts(wb)


if __name__ == '__main__':
    main()
